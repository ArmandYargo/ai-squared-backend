# func_inputsheet.py
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional, Dict, List
import random
import pandas as pd
from openpyxl import load_workbook
from datetime import date

# Reuse AI JSON helper
from func_classify_data import _llm_json

# ---------------------------
# Original column layouts
# ---------------------------
DEFAULT_COLUMNS_COMP_ATT = [
    "time_tbl","use_tbl","component","subcomponent","weight",
    "primary_failure_mode","pf_n","tf_beta","tf_eta",
    "mttr_fail","mttr_replace","mttr_repair",
    "cost_fail","cost_replace","cost_repair",
    "cb_ind","prev_ind","time_use","base_usage","factor_m",
    "deg_cond","insp_det_prob","cm_det_prob","cond_det_insp",
    "cond_det_cm","impr_rate","nserv_rate",
]

DEFAULT_COLUMNS_TIMELINE = [
    "from","to","maintenance_practice","rating",
    "serv_ind","pro_active_ind","active_ind",
]

DEFAULT_COLUMNS_USAGE = ["from","to","usage"]


def create_input_template(
    out_path: str | Path,
    *,
    n_timelines: int = 4,
    n_usages: int = 1,
    columns_comp_att: Optional[Iterable[str]] = None,
    columns_timeline: Optional[Iterable[str]] = None,
    columns_usage: Optional[Iterable[str]] = None,
    autosize: bool = True,
    overwrite: bool = True,
    freeze_header: bool = True,
) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.exists() and not overwrite:
        raise FileExistsError(f"{out_path} already exists and overwrite=False")

    comp_cols = list(columns_comp_att or DEFAULT_COLUMNS_COMP_ATT)
    tl_cols = list(columns_timeline or DEFAULT_COLUMNS_TIMELINE)
    us_cols = list(columns_usage or DEFAULT_COLUMNS_USAGE)

    dfs = {"comp_att": pd.DataFrame(columns=comp_cols)}
    for i in range(n_timelines):
        dfs[f"timeline_{i}"] = pd.DataFrame(columns=tl_cols)
    for i in range(n_usages):
        dfs[f"usage_{i}"] = pd.DataFrame(columns=us_cols)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in dfs.items():
            df.to_excel(writer, sheet_name=name, index=False)

    if autosize or freeze_header:
        _postprocess_workbook(out_path, autosize=autosize, freeze_header=freeze_header)

    return out_path


def _postprocess_workbook(xlsx_path: Path, *, autosize: bool, freeze_header: bool) -> None:
    wb = load_workbook(xlsx_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if freeze_header:
            ws.freeze_panes = "A2"

        if autosize:
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    val = cell.value
                    if val is None:
                        continue
                    try:
                        max_len = max(max_len, len(str(val)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max(10, max_len + 2)

    wb.save(xlsx_path)


# ===========================
# Build from summary data
# ===========================
MODEL_DEFAULT = "gpt-5.2"

# --------------------------------------------------------------------------------------
# NOTE: RAM_Simulation_Model_3.py matches maintenance_practice strings case-sensitively.
# Canonical values:
#   - Reactive
#   - Corrective
#   - Preventative
#   - Condition based
# This helper normalizes common variants produced by the wizard/template.

_MP_CANON = {
    "reactive": "Reactive",
    "corrective": "Corrective",
    "preventative": "Preventative",
    "preventive": "Preventative",
    "condition based": "Condition based",
    "condition_based": "Condition based",
    "conditionbased": "Condition based",
}


def normalize_maintenance_practice(value: object) -> object:
    """Return canonical maintenance practice string when possible; otherwise passthrough."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return value
    s = str(value).strip()
    key = s.lower().replace("_", " ")
    return _MP_CANON.get(key, s)

def _norm(s: str) -> str:
    s = (str(s) if s is not None else "").strip().lower()
    for ch in [" ", "-", "/"]:
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")

def _find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    """
    Find a column in df matching any candidate (case/space/underscore-insensitive).
    """
    norm_map = {col: _norm(col) for col in df.columns}
    cand_norm = [_norm(c) for c in candidates]
    # Exact normalized match
    for col, nc in norm_map.items():
        if nc in cand_norm:
            return col
    # Prefix match fallback
    for col, nc in norm_map.items():
        if any(nc.startswith(cn) for cn in cand_norm):
            return col
    return None

def _ai_decisions_for_components(
    components: List[str],
    primary_modes: List[Optional[str]],
    machine_type: str,
    *,
    model: str = MODEL_DEFAULT,
) -> pd.DataFrame:
    """
    Ask AI to propose: cb_ind (0/1), prev_ind (0/1),
    insp_det_prob, cm_det_prob, cond_det_insp, cond_det_cm in [0.1, 0.9].
    Always returns a DataFrame with a 'subcomponent' column aligned to `components`.
    """
    def _clip01(x: float) -> float:
        try:
            return float(min(0.9, max(0.1, x)))
        except Exception:
            return 0.5

    components = [str(c) if pd.notna(c) else "unknown" for c in components]
    primary_modes = [None if (pd.isna(m) or str(m).strip() == "") else str(m) for m in primary_modes]

    items = [{"subcomponent": c, "primary_failure_mode": (m or "")} for c, m in zip(components, primary_modes)]
    schema = {
        "name": "component_policy",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "items": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "subcomponent": {"type": "string"},
                            "cb_ind": {"type": "integer", "minimum": 0, "maximum": 1},
                            "prev_ind": {"type": "integer", "minimum": 0, "maximum": 1},
                            "insp_det_prob": {"type": "number"},
                            "cm_det_prob": {"type": "number"},
                            "cond_det_insp": {"type": "number"},
                            "cond_det_cm": {"type": "number"},
                        },
                        "required": ["subcomponent","cb_ind","prev_ind","insp_det_prob","cm_det_prob","cond_det_insp","cond_det_cm"],
                    }
                }
            },
            "required": ["items"],
            "additionalProperties": False,
        },
    }
    usr = {
        "role": "user",
        "content": (
            "You are a reliability engineer. For each subcomponent, decide:\n"
            "- cb_ind: 1 if condition-based monitoring is sensible/feasible; else 0\n"
            "- prev_ind: 1 if preventative maintenance is sensible; else 0\n"
            "- insp_det_prob: probability of detecting faults by visual inspection (0.1..0.9)\n"
            "- cm_det_prob: probability of detecting faults using condition monitoring (0.1..0.9)\n"
            "- cond_det_insp: replacement/repair decision threshold under inspection (0.1=almost broken .. 0.9=new)\n"
            "- cond_det_cm: same under condition monitoring (0.1..0.9)\n"
            "Make pragmatic, plant-ready choices. Consider the machine type and common practices.\n"
            f"Machine type: {machine_type}\n"
            f"Items: {items}\n"
            "Return ONLY JSON per the schema."
        ),
    }
    try:
        obj = _llm_json(
            model=model,
            messages=[
                {"role": "system", "content": "Be concise and realistic; clip outputs to 0.1..0.9."},
                usr,
            ],
            json_schema=schema,
            temperature=0.1,
            max_tokens=1200,
        )
        rows = []
        for it in obj.get("items", []):
            rows.append({
                "subcomponent": it.get("subcomponent",""),
                "cb_ind": int(it.get("cb_ind", 0)),
                "prev_ind": int(it.get("prev_ind", 1)),
                "insp_det_prob": _clip01(float(it.get("insp_det_prob", 0.5))),
                "cm_det_prob": _clip01(float(it.get("cm_det_prob", 0.5))),
                "cond_det_insp": _clip01(float(it.get("cond_det_insp", 0.5))),
                "cond_det_cm": _clip01(float(it.get("cond_det_cm", 0.5))),
            })
        ai_df = pd.DataFrame(rows)
    except Exception:
        ai_df = pd.DataFrame()

    if ai_df.empty or "subcomponent" not in ai_df.columns:
        ai_df = pd.DataFrame({
            "subcomponent": components,
            "cb_ind": [1]*len(components),
            "prev_ind": [1]*len(components),
            "insp_det_prob": [0.5]*len(components),
            "cm_det_prob": [0.6]*len(components),
            "cond_det_insp": [0.6]*len(components),
            "cond_det_cm": [0.7]*len(components),
        })

    ai_df = (
        ai_df.assign(_sub_key=ai_df["subcomponent"].map(_norm))
        .drop_duplicates("_sub_key", keep="first")
        .drop(columns="_sub_key")
        .reset_index(drop=True)
    )
    return ai_df


def build_comp_att_from_summary(
    summary_df: pd.DataFrame,
    *,
    machine_type: str,
    model: str = MODEL_DEFAULT,
) -> pd.DataFrame:
    """
    Map category_summary -> comp_att per your rules.
    Rows are strictly driven by UNIQUE Breakdown Categories (subcomponent).
    Robust to header variations (case/spacing/underscore).
    """
    df = summary_df.copy()

    col_category = _find_col(df, ["Breakdown Category", "breakdown_category", "category"])
    col_pfm = _find_col(df, ["Primary Failure Mode", "primary_failure_mode", "primary_failure_modes"])
    col_pfn = _find_col(df, ["pf_n", "pf n"])
    col_beta = _find_col(df, ["Beta"])
    col_eta  = _find_col(df, ["Eta"])
    col_mttrf = _find_col(df, ["MTTR_fail", "mttr_fail"])
    col_mttrrep = _find_col(df, ["MTTR_replace", "mttr_replace"])
    col_mttrr = _find_col(df, ["MTTR_repair", "mttr_repair"])
    col_cnt = _find_col(df, ["Downtime Count", "count", "downtime_count"])
    col_tot = _find_col(df, ["Total Downtime", "total_downtime", "hours_total"])  # optional

    missing = [name for name, col in {
        "Breakdown Category": col_category,
        "Primary Failure Mode": col_pfm,
        "pf_n": col_pfn,
        "Beta": col_beta,
        "Eta": col_eta,
        "MTTR_fail": col_mttrf,
        "MTTR_replace": col_mttrrep,
        "MTTR_repair": col_mttrr,
        "Downtime Count": col_cnt,
    }.items() if col is None]

    if missing:
        raise ValueError(f"category_summary missing required columns (robust check): {missing}")

    base = df[
        [col_category, col_pfm, col_pfn, col_beta, col_eta, col_mttrf, col_mttrrep, col_mttrr, col_cnt]
        + ([col_tot] if col_tot else [])
    ].copy()

    base = base.drop_duplicates(subset=[col_category]).reset_index(drop=True)

    src = pd.DataFrame({
        "subcomponent": base[col_category].astype(str),
        "primary_failure_mode": base[col_pfm],
        "pf_n": base[col_pfn],
        "tf_beta": base[col_beta],
        "tf_eta": base[col_eta],
        "mttr_fail": base[col_mttrf],
        "mttr_replace": base[col_mttrrep],
        "mttr_repair": base[col_mttrr],
        "count": base[col_cnt],
        "tot_hours": base[col_tot] if col_tot else 0.0,
    })

    n_rows = len(src)
    if n_rows == 0:
        return pd.DataFrame(columns=DEFAULT_COLUMNS_COMP_ATT)

    ai_df = _ai_decisions_for_components(
        components=src["subcomponent"].astype(str).tolist(),
        primary_modes=pd.Series(src["primary_failure_mode"]).astype(str).tolist(),
        machine_type=machine_type,
        model=model,
    )

    src["_sub_key"] = src["subcomponent"].map(_norm)
    ai_df["_sub_key"] = ai_df["subcomponent"].map(_norm)

    merged = src.merge(ai_df.drop_duplicates("_sub_key"), on="_sub_key", how="left", suffixes=("", "_ai"))
    merged["cb_ind"] = merged["cb_ind"].fillna(1).astype(int)
    merged["prev_ind"] = merged["prev_ind"].fillna(1).astype(int)
    for c, d in [
        ("insp_det_prob", 0.5),
        ("cm_det_prob", 0.6),
        ("cond_det_insp", 0.6),
        ("cond_det_cm", 0.7),
    ]:
        if c not in merged.columns:
            merged[c] = d
        else:
            merged[c] = merged[c].fillna(d)

    rng = random.Random(0)
    time_tbl = [rng.randint(0,3) for _ in range(n_rows)]
    use_tbl = [0]*n_rows
    component = [machine_type]*n_rows

    unique_cats = pd.Series(merged["subcomponent"]).astype(str).nunique(dropna=False)
    weight_val = 1.0 / float(unique_cats if unique_cats else 1)
    weight = [weight_val]*n_rows

    out = pd.DataFrame(columns=DEFAULT_COLUMNS_COMP_ATT)

    out["time_tbl"] = time_tbl
    out["use_tbl"] = use_tbl
    out["component"] = component
    out["subcomponent"] = merged["subcomponent"].astype(str)

    out["weight"] = weight
    out["primary_failure_mode"] = merged["primary_failure_mode"]
    out["pf_n"] = merged["pf_n"]
    out["tf_beta"] = merged["tf_beta"]
    out["tf_eta"] = merged["tf_eta"]

    out["mttr_fail"] = merged["mttr_fail"]
    out["mttr_replace"] = merged["mttr_replace"]
    out["mttr_repair"] = merged["mttr_repair"]

    out["cost_fail"] = "NA"
    out["cost_replace"] = "NA"
    out["cost_repair"] = "NA"

    out["cb_ind"] = merged["cb_ind"].astype(int)
    out["prev_ind"] = merged["prev_ind"].astype(int)

    out["time_use"] = 1
    out["base_usage"] = "NA"
    out["factor_m"] = "NA"

    pfm = (merged["primary_failure_mode"].astype(str).str.strip())

    # ---------- PATCH: columns that can receive "NA" must be object dtype ----------
    out["deg_cond"] = 0.1
    out["deg_cond"] = out["deg_cond"].astype("object")  # allow literal "NA"

    out["insp_det_prob"] = merged["insp_det_prob"]
    out["cm_det_prob"] = merged["cm_det_prob"]
    out["cond_det_insp"] = merged["cond_det_insp"]
    out["cond_det_cm"] = merged["cond_det_cm"]

    for _c in ["insp_det_prob", "cm_det_prob", "cond_det_insp", "cond_det_cm"]:
        out[_c] = out[_c].astype("object")  # allow literal "NA"
    # ---------------------------------------------------------------------------

    out.loc[pfm.eq("Random"), "deg_cond"] = "NA"

    mask_random = pfm.eq("Random")
    for col in ["insp_det_prob", "cm_det_prob", "cond_det_insp", "cond_det_cm"]:
        out.loc[mask_random, col] = "NA"

    mask_no_cb = out["cb_ind"].astype(int).eq(0)
    out.loc[mask_no_cb, "cm_det_prob"] = "NA"
    out.loc[mask_no_cb, "cond_det_insp"] = "NA"

    out["impr_rate"] = "NA"
    out["nserv_rate"] = 10

    out = out.sort_values("subcomponent").reset_index(drop=True)
    return out


def build_timelines(start_date: date) -> Dict[str, pd.DataFrame]:
    ts = pd.to_datetime(start_date)
    to = ts + pd.DateOffset(years=50)

    def _row(mp: str, rating: float, serv: int, pro: int, active: int):
        return pd.DataFrame([{
            "from": ts.date(),
            "to": to.date(),
            "maintenance_practice": mp,
            "rating": rating,
            "serv_ind": serv,
            "pro_active_ind": pro,
            "active_ind": active,
        }], columns=DEFAULT_COLUMNS_TIMELINE)

    return {
        "timeline_0": _row("Reactive", 0.0, 0, 0, 1),
        "timeline_1": _row("Corrective", 0.0, 0, 0, 1),
        "timeline_2": _row("Preventative", 0.5, 1, 0, 1),
        "timeline_3": _row("Condition based", 0.8, 1, 0, 1),
    }


def build_usage0() -> pd.DataFrame:
    return pd.DataFrame([{"from": "NA", "to": "NA", "usage": "NA"}], columns=DEFAULT_COLUMNS_USAGE)


def export_input_workbook(
    out_path: str | Path,
    *,
    comp_att: pd.DataFrame,
    timelines: Dict[str, pd.DataFrame],
    usage0: pd.DataFrame,
    autosize: bool = True,
    overwrite: bool = True,
) -> Path:
    out_path = Path(out_path)
    if out_path.exists() and not overwrite:
        raise FileExistsError(f"{out_path} already exists and overwrite=False")

    # Normalize maintenance_practice strings to what the RAM model expects.
    # This prevents fragile, case-sensitive failures in RAM_Simulation_Model_3.
    fixed_timelines: Dict[str, pd.DataFrame] = {}
    for name, df in timelines.items():
        tdf = df.copy()
        if "maintenance_practice" in tdf.columns:
            tdf["maintenance_practice"] = tdf["maintenance_practice"].map(normalize_maintenance_practice)
        fixed_timelines[name] = tdf

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        comp_att.to_excel(writer, sheet_name="comp_att", index=False)
        for name, df in fixed_timelines.items():
            df.to_excel(writer, sheet_name=name, index=False)
        usage0.to_excel(writer, sheet_name="usage_0", index=False)

    _postprocess_workbook(out_path, autosize=True, freeze_header=True)
    return out_path
