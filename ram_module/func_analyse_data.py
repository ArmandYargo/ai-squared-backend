# func_analyse_data.py
from __future__ import annotations
from pathlib import Path
from typing import Union, Dict, List, Optional
import pandas as pd
from math import gamma

# Reuse JSON helper for AI scoring of Beta (complexity) when needed
from func_classify_data import _llm_json

MODEL_DEFAULT = "gpt-5.2"

# --- Primary Failure Mode -> pf_n mapping ---
_PF_VALUE_MAP: Dict[str, float] = {
    "Wear": 1.0,
    "Random": 1.0,
    "Fatigue": 0.33,
    "Corrosion": 1.8,
}

def summarise_breakdowns(
    df: pd.DataFrame,
    *,
    category_col: str = "breakdown_category",
    duration_col: str = "breakdown_duration_hours",
) -> pd.DataFrame:
    """
    Build a summary table with:
      - Breakdown Category (unique)
      - Downtime Count (number of rows per category)
      - Total Downtime (sum of duration_col per category)  [hours]
    """
    if category_col not in df.columns:
        raise ValueError(f"Expected column '{category_col}' in dataframe.")

    data = df.copy()
    data[category_col] = data[category_col].fillna("unknown").astype(str)

    if duration_col in data.columns:
        hours = pd.to_numeric(data[duration_col], errors="coerce").fillna(0.0)
    else:
        hours = pd.Series(0.0, index=data.index)

    counts = data.groupby(category_col, dropna=False).size().rename("Downtime Count")
    totals = hours.groupby(data[category_col], dropna=False).sum().rename("Total Downtime")  # hours

    summary = pd.concat([counts, totals], axis=1).reset_index()
    summary = summary.rename(columns={category_col: "Breakdown Category"})
    summary = summary.sort_values(
        by=["Total Downtime", "Downtime Count"], ascending=[False, False], kind="mergesort"
    ).reset_index(drop=True)
    return summary


# -------------------- New helpers for MTTR/MTBF/Beta/Eta --------------------

# Heuristic baseline for complexity (1–3) by category token
_BETA_BASELINE = {
    "drive": 3,
    "gearbox": 3,
    "motor": 3,
    "pulley": 2,
    "take_up": 2,
    "belt": 2,
    "chute": 2,
    "electrical_control": 2,
    "sensor": 1,
    "control_instrumentation": 1,  # "C&I"
    "idler": 1,
    "structure": 1,
    "lubrication": 1,
    "guarding": 1,
    "spillage": 1,
    "alignment": 1,
    "other": 2,
}

def _normalize_cat(s: str) -> str:
    s = (s or "").lower().strip()
    for ch in ["/", "-", " "]:
        s = s.replace(ch, "_")
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_") or "other"

def _ai_score_beta_for_unknown(
    categories: List[str],
    *,
    model: str = MODEL_DEFAULT,
) -> Dict[str, int]:
    """
    Ask AI to score complexity 1..3 for categories not in the baseline map.
    Returns dict of normalized_category -> int in {1,2,3}.
    """
    if not categories:
        return {}

    schema = {
        "name": "beta_scores",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "scores": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {"category": {"type": "string"}, "beta": {"type": "integer", "minimum": 1, "maximum": 3}},
                        "required": ["category", "beta"],
                    },
                }
            },
            "required": ["scores"],
            "additionalProperties": False,
        },
    }
    prompt = {
        "role": "user",
        "content": (
            "Score the fault-finding complexity of each breakdown category on a 1–3 scale "
            "(1 = least complex, 3 = most complex). Consider breadth of components, diagnosis steps, "
            "and typical troubleshooting difficulty.\n"
            f"Categories: {categories}\n"
            "Return ONLY JSON with key 'scores', each item: {category, beta}."
        ),
    }
    try:
        obj = _llm_json(
            model=model,
            messages=[
                {"role": "system", "content": "You are a reliability engineer assigning complexity scores."},
                prompt,
            ],
            json_schema=schema,
            temperature=0.0,
            max_tokens=800,
        )
        out: Dict[str, int] = {}
        for item in obj.get("scores", []):
            cat = _normalize_cat(str(item.get("category", "")))
            b = int(item.get("beta", 2))
            b = 1 if b < 1 else 3 if b > 3 else b
            out[cat] = b
        return out
    except Exception:
        # If AI call fails, default to 2
        return { _normalize_cat(c): 2 for c in categories }

def _get_beta_series(categories: pd.Series, *, model: str = MODEL_DEFAULT) -> pd.Series:
    """
    Map category -> Beta (1..3). Use baseline map first, then AI for unknowns, fallback=2.
    """
    norm = categories.fillna("other").astype(str).map(_normalize_cat)
    uniq = sorted(set(norm))
    # split known vs unknown
    known = {c: _BETA_BASELINE[c] for c in uniq if c in _BETA_BASELINE}
    unknown = [c for c in uniq if c not in known]
    if unknown:
        ai_map = _ai_score_beta_for_unknown(unknown, model=model)
        known.update({k: ai_map.get(k, 2) for k in unknown})
    return norm.map(lambda c: known.get(c, 2))


def enrich_summary_with_mttr_mtbf(
    summary_df: pd.DataFrame,
    *,
    days_in_range: Optional[int],
    model: str = MODEL_DEFAULT,
) -> pd.DataFrame:
    """
    Add per-row metrics:
      - MTTR_repair   = ((Total Downtime / 24) / Count) * 0.1
      - MTTR_replace  =  (Total Downtime / 24) / Count
      - MTTR_fail     = ((Total Downtime / 24) / Count) * 10
      - MTBF          = Days in range / Count
      - Beta          = AI complexity score 1..3 for the category
      - Eta           = MTBF / Gamma(1 + (1/Beta))

    Notes:
    - 'Total Downtime' is in hours (by construction).
    - If days_in_range is None, MTBF and Eta will be NaN.
    """
    df = summary_df.copy()

    # Ensure numeric
    cnt = pd.to_numeric(df["Downtime Count"], errors="coerce")
    tot_h = pd.to_numeric(df["Total Downtime"], errors="coerce")
    cnt_safe = cnt.where(cnt > 0)

    per_event_days = (tot_h / 24.0) / cnt_safe
    df["MTTR_repair"] = per_event_days * 0.1
    df["MTTR_replace"] = per_event_days
    df["MTTR_fail"] = per_event_days * 10.0

    if days_in_range is not None and days_in_range > 0:
        df["MTBF"] = days_in_range / cnt_safe
    else:
        df["MTBF"] = pd.NA

    # Beta via baseline + AI
    df["Beta"] = _get_beta_series(df["Breakdown Category"], model=model)

    # Eta = MTBF / Gamma(1 + 1/Beta)
    def _eta_row(row):
        try:
            b = float(row["Beta"])
            mtbf = float(row["MTBF"])
            if b <= 0 or pd.isna(mtbf):
                return pd.NA
            return mtbf / gamma(1.0 + (1.0 / b))
        except Exception:
            return pd.NA

    df["Eta"] = df.apply(_eta_row, axis=1)

    return df


# -------- Primary Failure Mode + pf_n (per category) --------
def _primary_failure_mode_table(
    classified_df: pd.DataFrame,
    *,
    category_col: str = "breakdown_category",
    pfm_col_preferred: str = "primary_failure_modes",
) -> pd.DataFrame:
    """
    Determine, per breakdown category:
      - Primary Failure Mode (the most frequent of Wear/Random/Fatigue/Corrosion)
      - pf_n (mapped value: Wear=1, Random=1, Fatigue=0.33, Corrosion=1.8)
    Returns a table with columns: ['Breakdown Category','Primary Failure Mode','pf_n']
    """
    pfm_col = pfm_col_preferred if pfm_col_preferred in classified_df.columns else (
        "failure_modes" if "failure_modes" in classified_df.columns else None
    )
    if pfm_col is None:
        # Nothing to compute
        return pd.DataFrame(columns=["Breakdown Category", "Primary Failure Mode", "pf_n"])

    temp = classified_df[[category_col, pfm_col]].copy()
    temp[pfm_col] = temp[pfm_col].astype(str)

    # Count occurrences per category x mode
    counts = (
        temp.groupby([category_col, pfm_col], dropna=False)
            .size()
            .reset_index(name="cnt")
    )

    # Prefer only the 4 known modes; if others appear, we keep them but they won't get pf_n
    counts["pf_n"] = counts[pfm_col].map(_PF_VALUE_MAP)
    counts["is_known_mode"] = counts[pfm_col].isin(_PF_VALUE_MAP.keys())

    # Sort so that the "winner" per category is first:
    # - highest count
    # - prefer known modes (True before False)
    # - if still tied, higher pf_n
    # - final tie-break: alphabetical mode (stable)
    counts = counts.sort_values(
        by=[category_col, "cnt", "is_known_mode", "pf_n", pfm_col],
        ascending=[True, False, False, False, True],
        kind="mergesort",
    )

    # Pick top row per category
    winners = counts.drop_duplicates(subset=[category_col], keep="first").copy()
    winners = winners.rename(columns={
        category_col: "Breakdown Category",
        pfm_col: "Primary Failure Mode",
    })[["Breakdown Category", "Primary Failure Mode", "pf_n"]]

    return winners


def _merge_primary_failure_mode(summary_df: pd.DataFrame, classified_df: pd.DataFrame) -> pd.DataFrame:
    winners = _primary_failure_mode_table(classified_df)
    if winners.empty:
        out = summary_df.copy()
        out["Primary Failure Mode"] = pd.NA
        out["pf_n"] = pd.NA
        return out
    out = summary_df.merge(winners, on="Breakdown Category", how="left")
    return out


def append_summary_sheet_to_excel(
    summary_df: pd.DataFrame,
    excel_path: Union[str, Path],
    *,
    sheet_name: str = "category_summary",
    autosize: bool = True,
) -> None:
    """Append/replace a sheet in an existing Excel file with the summary table."""
    excel_path = Path(excel_path)
    
    # Build ExcelWriter params based on whether file exists
    writer_params = {
        "engine": "openpyxl",
    }
    
    if excel_path.exists():
        writer_params["mode"] = "a"
        writer_params["if_sheet_exists"] = "replace"
    else:
        writer_params["mode"] = "w"
    
    with pd.ExcelWriter(excel_path, **writer_params) as writer:
        summary_df.to_excel(writer, sheet_name=sheet_name, index=False)

    if autosize:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for col in ws.columns:
                    max_len = 0
                    for cell in col:
                        val = "" if cell.value is None else str(cell.value)
                        if len(val) > max_len:
                            max_len = len(val)
                    width = max(10, min(max_len + 2, 60))
                    ws.column_dimensions[col[0].column_letter].width = width
            wb.save(excel_path)
        except Exception:
            pass  # non-fatal


def analyse_and_append_to_excel(
    classified_df: pd.DataFrame,
    excel_path: Union[str, Path],
    *,
    category_col: str = "breakdown_category",
    duration_col: str = "breakdown_duration_hours",
    sheet_name: str = "category_summary",
    days_in_range: Optional[int] = None,
    model: str = MODEL_DEFAULT,
) -> pd.DataFrame:
    """
    Build the summary, add:
      - Primary Failure Mode & pf_n (mode per category with mapped value)
      - MTTR_repair / MTTR_replace / MTTR_fail
      - MTBF / Beta / Eta
    and write as a new/updated sheet.
    - If days_in_range is provided, MTBF and Eta are computed; otherwise they are left blank (NA).
    """
    base = summarise_breakdowns(classified_df, category_col=category_col, duration_col=duration_col)
    # add Primary Failure Mode + pf_n
    base_with_pfm = _merge_primary_failure_mode(base, classified_df)
    # enrich with MTTR/MTBF/Beta/Eta
    enriched = enrich_summary_with_mttr_mtbf(base_with_pfm, days_in_range=days_in_range, model=model)
    append_summary_sheet_to_excel(enriched, excel_path, sheet_name=sheet_name)
    return enriched

