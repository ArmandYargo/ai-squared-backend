# func_date_filter.py
from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional, Tuple, Union
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook

# Reuse existing helpers (no behavior change to current modules)
from func_classify_data import _llm_json, _snake, ai_build_machine_filter


MODEL_DEFAULT = "gpt-5.2"


# ---------- PATCH: robust datetime parsing to avoid "Could not infer format" spam ----------
import warnings

def _to_datetime_series(s: pd.Series, *, dayfirst: bool) -> pd.Series:
    """
    Parse dates robustly without noisy 'Could not infer format' warnings.
    - Uses pandas format='mixed' when available (pandas>=2.0)
    - Falls back to a two-pass parse for older pandas
    """
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Could not infer format")
        try:
            # pandas>=2.0
            return pd.to_datetime(s, errors="coerce", dayfirst=dayfirst, utc=False, format="mixed")
        except TypeError:
            # pandas<2.0
            p1 = pd.to_datetime(s, errors="coerce", utc=False)
            if getattr(p1, "isna", None) is not None and p1.isna().mean() > 0.2:
                p2 = pd.to_datetime(s, errors="coerce", dayfirst=dayfirst, utc=False)
                p1 = p1.fillna(p2)
            return p1


# ---------- internal: pick a date-like START column (LLM + fallback) ----------
def _ai_pick_start_date_column(
    df: pd.DataFrame,
    *,
    model: str = MODEL_DEFAULT,
    dayfirst: bool = True,
) -> Optional[str]:
    """
    Identify the single column most likely to represent a 'start' date/time for an event.
    Prefers names like 'Notification Date', 'Malfunction Start', 'Start', 'Reported', etc.
    Combines LLM guidance with heuristic fallback and plausibility (parseable as datetime).
    """
    HINTS = [
        "notification date", "malfunction start", "malfunct start", "start date", "start",
        "date reported", "reported on", "created on", "created date", "created",
        "breakdown start", "actual start", "start_time", "time start", "failure start",
        "start of malfunction", "fault start", "start timestamp", "began", "opened on",
    ]

    def _name_score(col: str) -> float:
        cn = _snake(col)
        best = 0.0
        for h in HINTS:
            r = pd.get_option("display.precision")  # tiny noop to avoid lint on unused import
            # simple difflib not imported here; approximate using token membership
            # (we already lean on LLM; this is just a light bias)
            if _snake(h) in cn:
                best = max(best, 0.8)
        # gentle boost if it *looks* like a date-ish name
        for tok in ("date", "start", "reported", "created", "open", "began", "malfunct"):
            if tok in cn:
                best = max(best, 0.6)
        return min(best, 1.0)

    def _parseable_ratio(s: pd.Series) -> float:
        if s.notna().sum() == 0:
            return 0.0
        # PATCH: use robust parser (suppresses warning)
        parsed = _to_datetime_series(s, dayfirst=dayfirst)
        return float(parsed.notna().mean())

    cols = list(df.columns)[:60]
    summaries = []
    for c in cols:
        s = df[c]
        nn = int(s.notna().sum())
        if nn == 0:
            pr = 0.0
            ex = []
        else:
            pr = _parseable_ratio(s)
            ex = [str(v)[:40] for v in s.dropna().astype(str).head(5).tolist()]
        summaries.append(
            {"name": c, "snake": _snake(c), "dtype": str(s.dtype), "non_null": nn, "parseable_ratio": round(pr, 3), "examples": ex}
        )

    # LLM suggestion (structured)
    schema = {
        "name": "pick_start_date",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "start_date_column": {"type": ["string", "null"]},
                "rationale": {"type": "string"},
            },
            "required": ["start_date_column"],
            "additionalProperties": False,
        },
    }
    try:
        obj = _llm_json(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You select the single column that represents the START date/time of a maintenance event. "
                        "Prefer names like 'Notification Date', 'Malfunction Start', 'Start', 'Reported', 'Created'. "
                        "Avoid end/complete/finish/closed timestamps, IDs, and categorical fields."
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        "Columns summary (first 60):\n"
                        f"{summaries}\n\n"
                        "Return ONLY JSON with 'start_date_column' (exact name) or null if none."
                    ),
                },
            ],
            json_schema=schema,
            temperature=0.0,
            max_tokens=600,
        )
        candidate = obj.get("start_date_column")
    except Exception:
        candidate = None

    # Heuristic fallback
    if candidate not in df.columns:
        best_col, best_score = None, -1.0
        for c in cols:
            score = 0.65 * _name_score(c) + 0.35 * _parseable_ratio(df[c])
            if score > best_score:
                best_col, best_score = c, score
        if best_col is not None:
            # require some minimum plausibility
            if _parseable_ratio(df[best_col]) >= 0.25 or _name_score(best_col) >= 0.7:
                candidate = best_col

    # Final plausibility check
    if candidate is None:
        return None
    try:
        # PATCH: robust parser (suppresses warning)
        parsed = _to_datetime_series(df[candidate], dayfirst=dayfirst)
        if parsed.notna().sum() == 0:
            return None
    except Exception:
        return None
    return candidate


# ---------- internal: parse natural-language date range via LLM (inclusive) ----------
def _ai_parse_date_range(
    user_text: str,
    *,
    today: Optional[date] = None,
    model: str = MODEL_DEFAULT,
) -> Tuple[date, date, Dict]:
    """
    Interpret a free-form date range like 'from 2025/03/12 to 2025/06/02' or 'November to Jan'.
    Returns inclusive (start_date, end_date) as Python dates.
    """
    today = today or date.today()
    schema = {
        "name": "date_range",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "start": {"type": "string", "description": "YYYY-MM-DD"},
                "end": {"type": "string", "description": "YYYY-MM-DD"},
                "notes": {"type": "string"},
            },
            "required": ["start", "end"],
            "additionalProperties": False,
        },
    }
    obj = _llm_json(
        model=model,
        messages=[
            {
                "role": "system",
                "content": (
                    "You interpret natural-language date ranges and return concrete calendar dates. "
                    "Assume the user wants inclusive ranges. If month names lack years, infer the most recent sensible span relative to 'today'."
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Today is {today.isoformat()}.\n"
                    f"Interpret this as an inclusive date range and return ONLY JSON with 'start' and 'end' (YYYY-MM-DD):\n"
                    f"\"{user_text}\""
                ),
            },
        ],
        json_schema=schema,
        temperature=0.0,
        max_tokens=300,
    )
    try:
        sd = date.fromisoformat(obj["start"])
        ed = date.fromisoformat(obj["end"])
    except Exception as e:
        raise ValueError(f"Could not parse date range from: {user_text!r}. Error: {e}")

    # normalize ordering
    if sd > ed:
        sd, ed = ed, sd
    return sd, ed, {"raw": user_text, "today": today.isoformat(), "model_notes": obj.get("notes")}


# ---------- public: extract dates after machine selection + stamp summary ----------
def extract_dates_and_update_output(
    df_master: pd.DataFrame,
    mapping: Dict,
    machine_query: str,
    user_date_range_text: str,
    out_excel_path: Union[str, Path],
    *,
    model: str = MODEL_DEFAULT,
    match_scope: str = "both",  # 'floc' | 'desc' | 'both'
    dayfirst: bool = True,
    summary_sheet: str = "category_summary",
) -> Dict:
    """
    New post-machine step that:
      1) Detects the best START date column on the machine-filtered subset ONLY.
      2) Interprets a natural-language date range provided by the user (inclusive).
      3) Filters the machine subset again by that date range (without exposing the date column in 'classified').
      4) Writes the date-range span (days & hours) at the bottom of the existing 'category_summary' sheet.
         (Does not modify the 'classified' sheet.)
    """
    floc_col = mapping.get("functional_location")
    desc_col = mapping.get("description")
    if not floc_col or not desc_col or floc_col not in df_master.columns or desc_col not in df_master.columns:
        raise ValueError(f"Invalid mapping: {mapping}")

    # -------- 1) Filter to the chosen machine --------
    inc_re, exc_re, meta = ai_build_machine_filter(df_master[floc_col], machine_query, model=model)

    floc_s = df_master[floc_col].astype(str)
    desc_s = df_master[desc_col].astype(str)

    m_inc_floc = floc_s.str.contains(inc_re, na=False)
    m_inc_desc = desc_s.str.contains(inc_re, na=False)
    m_exc_floc = floc_s.str.contains(exc_re, na=False) if exc_re else pd.Series(False, index=df_master.index)
    m_exc_desc = desc_s.str.contains(exc_re, na=False) if exc_re else pd.Series(False, index=df_master.index)

    if match_scope == "floc":
        machine_mask = m_inc_floc & ~m_exc_floc
    elif match_scope == "desc":
        machine_mask = m_inc_desc & ~m_exc_desc
    else:
        machine_mask = (m_inc_floc | m_inc_desc) & ~(m_exc_floc | m_exc_desc)

    machine_subset = df_master[machine_mask].copy()
    n_machine = int(len(machine_subset))

    # -------- 2) Detect date column *on the machine subset only* --------
    date_col = _ai_pick_start_date_column(machine_subset, model=model, dayfirst=dayfirst)

    # Build a parsed datetime Series (may be empty if date_col is None)
    if date_col:
        # PATCH: robust parser (suppresses warning)
        date_series = _to_datetime_series(machine_subset[date_col], dayfirst=dayfirst).dt.date
    else:
        date_series = pd.Series([None] * n_machine, index=machine_subset.index)

    # -------- 3) Interpret user's natural-language date range (inclusive) --------
    start_d, end_d, parse_meta = _ai_parse_date_range(user_date_range_text, model=model)

    # Range stats (inclusive)
    days_inclusive = (end_d - start_d).days + 1
    if days_inclusive < 0:
        days_inclusive = 0
    hours_inclusive = days_inclusive * 24

    # -------- 4) Date filter --------
    if date_col:
        mask_range = (date_series >= start_d) & (date_series <= end_d)
        filtered_subset = machine_subset[mask_range].copy()
        n_date_filtered = int(len(filtered_subset))
    else:
        n_date_filtered = 0

    # -------- 5) Stamp the range info at the end of 'category_summary' sheet --------
    out_excel_path = Path(out_excel_path)
    try:
        wb = load_workbook(out_excel_path)
        if summary_sheet in wb.sheetnames:
            ws = wb[summary_sheet]
            row = ws.max_row + 2  # spacer row
            ws.cell(row=row, column=1, value="Date range (inclusive)")
            ws.cell(row=row, column=2, value=f"{start_d.isoformat()} to {end_d.isoformat()}")
            row += 1
            ws.cell(row=row, column=1, value="Days in range")
            ws.cell(row=row, column=2, value=days_inclusive)
            row += 1
            ws.cell(row=row, column=1, value="Hours in range (24h/day)")
            ws.cell(row=row, column=2, value=hours_inclusive)
            try:
                ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 10, 26)  # type: ignore[attr-defined]
                ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 10, 22)  # type: ignore[attr-defined]
            except Exception:
                pass
            wb.save(out_excel_path)
        else:
            wb.close()
    except Exception:
        pass

    return {
        "date_column": date_col,
        "start_date": start_d.isoformat(),
        "end_date": end_d.isoformat(),
        "days_inclusive": int(days_inclusive),
        "hours_inclusive": int(hours_inclusive),
        "n_machine_rows": n_machine,
        "n_date_filtered_rows": int(n_date_filtered),
        "notes": {"machine_regex": {"include": str(inc_re.pattern) if inc_re else None,
                                    "exclude": str(exc_re.pattern) if exc_re else None},
                  "parser": parse_meta},
    }
